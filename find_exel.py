import openpyxl as pd

def serchToCity(n):
    book = pd.open("City_Price.xlsx", read_only=True)
    sheet = book.worksheets[0]
    flag = False
    a = 1
    str = input("введите город: " ).lower().strip()
    for i in range(1, sheet.max_row +1):
        a = sheet[i][0].value.lower().strip()
        if a == str:
            a =  sheet[i][1].value
            flag = True
            break
    
    return a

def serchToMoscow(n):
    book = pd.open("City_Price.xlsx", read_only=True)
    sheet_2 = book.worksheets[1]
    for i in range(2, sheet_2.max_row +1):
        if n <= int(sheet_2[i][0].value):
            return sheet_2[i][1].value
# print (serchToMoscow(n))


def listOfCity():
    book = pd.open("City_Price.xlsx", read_only=True)
    sheet= book.worksheets[0]
    for i in range(2, sheet.max_row +1):
        print(sheet[i][0].value)
        
listOfCity()
